#!/usr/bin/env python3
"""Download and convert UCI Online Retail to a DuckDB-friendly CSV.

The source is the public UCI Machine Learning Repository dataset 352. The
dataset is not bundled; this script records the fetched archive hash and source
metadata so learners can preserve lineage without pretending the synthetic
course seeds are real records.
"""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime, timezone
import hashlib
import json
from pathlib import Path
import shutil
import urllib.request
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "data" / "uci_online_retail"
ARCHIVE = OUTPUT_DIR / "online_retail.zip"
WORKBOOK = OUTPUT_DIR / "Online Retail.xlsx"
CSV_PATH = OUTPUT_DIR / "online_retail.csv"
SOURCE_URL = "https://archive.ics.uci.edu/static/public/352/online+retail.zip"
DATASET_PAGE = "https://archive.ics.uci.edu/dataset/352/online-retail"
DOI = "10.24432/C5BW33"
LICENSE = "CC BY 4.0"
HEADERS = (
    "invoice_no", "stock_code", "description", "quantity", "invoice_date",
    "unit_price_gbp", "customer_id", "country",
)


def download() -> None:
    request = urllib.request.Request(
        SOURCE_URL,
        headers={"User-Agent": "dbt-engineering-course/1.0 (educational importer)"},
    )
    print(f"Downloading {SOURCE_URL}")
    with urllib.request.urlopen(request, timeout=120) as response, ARCHIVE.open("wb") as output:
        shutil.copyfileobj(response, output)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_workbook() -> None:
    with zipfile.ZipFile(ARCHIVE) as archive:
        members = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
        if len(members) != 1:
            raise RuntimeError(f"Expected one .xlsx file in UCI archive, found {members}")
        with archive.open(members[0]) as source, WORKBOOK.open("wb") as destination:
            shutil.copyfileobj(source, destination)


def csv_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else value


def convert(max_rows: int | None) -> int:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    source_headers = next(rows)
    if len(source_headers) != len(HEADERS):
        raise RuntimeError(f"Unexpected UCI workbook columns: {source_headers}")

    count = 0
    with CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        for source_row in rows:
            writer.writerow([csv_value(value) for value in source_row])
            count += 1
            if max_rows and count >= max_rows:
                break
    workbook.close()
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--max-rows",
        type=int,
        help="Optional positive row limit for a faster local sample; default imports every row.",
    )
    parser.add_argument("--refresh", action="store_true", help="Download the UCI archive again.")
    args = parser.parse_args()
    if args.max_rows is not None and args.max_rows <= 0:
        parser.error("--max-rows must be positive")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if args.refresh or not ARCHIVE.exists():
        download()
    else:
        print(f"Using cached {ARCHIVE.relative_to(ROOT)}")
    extract_workbook()
    row_count = convert(args.max_rows)
    metadata = {
        "title": "Online Retail",
        "creator": "Daqing Chen",
        "repository": "UCI Machine Learning Repository",
        "dataset_page": DATASET_PAGE,
        "download_url": SOURCE_URL,
        "doi": DOI,
        "license": LICENSE,
        "license_url": "https://creativecommons.org/licenses/by/4.0/",
        "retrieved_at_utc": datetime.now(timezone.utc).isoformat(),
        "archive_sha256": sha256(ARCHIVE),
        "rows_written": row_count,
        "row_limit": args.max_rows,
        "notes": "Raw public transactions converted from XLSX to CSV; no synthetic rows added.",
    }
    provenance = OUTPUT_DIR / "provenance.json"
    provenance.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {row_count:,} real transaction rows to {CSV_PATH.relative_to(ROOT)}")
    print(f"Recorded provenance in {provenance.relative_to(ROOT)}")
    print("Enable the optional models with: dbt build --vars '{\"enable_uci\": true}' --select path:models/real_data")


if __name__ == "__main__":
    main()

