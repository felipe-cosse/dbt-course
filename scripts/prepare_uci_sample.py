#!/usr/bin/env python3
"""Create a deterministic, redistributable learning sample from UCI Online Retail.

The source dataset is CC BY 4.0. Download `Online Retail.xlsx` from
https://archive.ics.uci.edu/dataset/352/online-retail, then run:

    python scripts/prepare_uci_sample.py "Online Retail.xlsx" \
      data/real/uci_online_retail_sample.csv

The script uses reservoir sampling so the sample spans the source workbook
without holding all 541k rows in memory.
"""

from __future__ import annotations

import argparse
import csv
import random
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADERS = [
    "invoice_no",
    "stock_code",
    "description",
    "quantity",
    "invoice_date",
    "unit_price_gbp",
    "customer_id",
    "country",
]


def normalized_row(values: tuple[Any, ...]) -> list[Any]:
    row = list(values[:8])
    if isinstance(row[4], datetime):
        row[4] = row[4].isoformat(sep=" ")
    if row[6] is not None:
        row[6] = str(int(row[6]))
    return row


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--rows", type=int, default=750)
    args = parser.parse_args()

    randomizer = random.Random(352)
    reservoir: list[list[Any]] = []
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    worksheet = workbook.active

    for source_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        row = normalized_row(values)
        if source_index <= args.rows:
            reservoir.append(row)
            continue
        replacement_index = randomizer.randint(1, source_index)
        if replacement_index <= args.rows:
            reservoir[replacement_index - 1] = row

    reservoir.sort(key=lambda row: (str(row[4]), str(row[0]), str(row[1])))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(reservoir)

    print(f"Wrote {len(reservoir)} real transaction rows to {args.output}")


if __name__ == "__main__":
    main()
